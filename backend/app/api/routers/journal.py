from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import Response
from sqlalchemy.orm import Session, joinedload
from datetime import date
from io import BytesIO
import csv
from ...dependencies import get_db
from ...application.dtos import JournalEntryIn, JournalEntryOut, JournalEntryDetailOut, EntryLineOut
from ...domain.models import JournalEntry, EntryLine, Account, Period, User
from ...infrastructure.unit_of_work import UnitOfWork
from ...security.auth import get_current_user
from typing import List, Dict, Union, Any

router = APIRouter(prefix="/journal", tags=["journal"])

@router.get("/entries")
def list_entries(
    company_id: int = Query(..., description="ID de la empresa"),
    period_id: int | None = Query(default=None, description="Filtrar por periodo"),
    date_from: date | None = Query(default=None, description="Fecha desde (YYYY-MM-DD)"),
    date_to: date | None = Query(default=None, description="Fecha hasta (YYYY-MM-DD)"),
    account_code: str | None = Query(default=None, description="Filtrar por código de cuenta"),
    status: str | None = Query(default=None, description="Filtrar por estado (POSTED, VOIDED)"),
    include_lines: bool = Query(default=False, description="Incluir líneas de cada asiento"),
    db: Session = Depends(get_db),
):
    # Configurar eager loading si se solicitan las líneas
    if include_lines:
        query = db.query(JournalEntry).options(
            joinedload(JournalEntry.lines).joinedload(EntryLine.account)
        ).filter(JournalEntry.company_id == company_id)
    else:
        query = db.query(JournalEntry).filter(JournalEntry.company_id == company_id)
    
    if period_id:
        query = query.filter(JournalEntry.period_id == period_id)
    if date_from:
        query = query.filter(JournalEntry.date >= date_from)
    if date_to:
        query = query.filter(JournalEntry.date <= date_to)
    if status:
        query = query.filter(JournalEntry.status == status)
    
    # Filtrar por cuenta si se especifica
    if account_code:
        query = query.join(EntryLine).join(Account).filter(Account.code == account_code).distinct()
    
    # Ordenar por fecha y número de asiento (ID) para el libro diario
    entries = query.order_by(JournalEntry.date, JournalEntry.id).all()
    
    result = []
    for entry in entries:
        total_debit = float(sum([float(l.debit) for l in entry.lines]))
        total_credit = float(sum([float(l.credit) for l in entry.lines]))
        
        if include_lines:
            # Obtener periodo
            period = db.query(Period).filter(Period.id == entry.period_id).first() if entry.period_id else None
            
            # Construir líneas con información de cuenta
            lines_out = []
            for line in entry.lines:
                account = line.account if hasattr(line, 'account') and line.account else db.query(Account).filter(Account.id == line.account_id).first()
                lines_out.append(EntryLineOut(
                    id=line.id,
                    account_code=account.code if account else "",
                    account_name=account.name if account else "",
                    debit=float(line.debit),
                    credit=float(line.credit),
                    memo=line.memo,
                    third_party_id=line.third_party_id,
                    cost_center=line.cost_center,
                ))
            
            result.append(JournalEntryDetailOut(
                id=entry.id,
                company_id=entry.company_id,
                date=entry.date,
                glosa=entry.glosa or "",
                currency=entry.currency,
                origin=entry.origin,
                status=entry.status,
                total_debit=total_debit,
                total_credit=total_credit,
                period_id=entry.period_id if entry.period_id else 0,
                period_year=period.year if period else 0,
                period_month=period.month if period else 0,
                exchange_rate=float(entry.exchange_rate),
                lines=lines_out,
                correlative=entry.correlative,
                reversed_entry_id=entry.reversed_entry_id,
                motor_metadata=entry.motor_metadata,
            ))
        else:
            result.append(JournalEntryOut(
                id=entry.id,
                company_id=entry.company_id,
                date=entry.date,
                glosa=entry.glosa or "",
                currency=entry.currency,
                origin=entry.origin,
                status=entry.status,
                total_debit=total_debit,
                total_credit=total_credit,
                period_id=entry.period_id if entry.period_id else None,
                correlative=entry.correlative,
                reversed_entry_id=entry.reversed_entry_id,
                motor_metadata=entry.motor_metadata,
            ))
    return result

def _user_display_name(user: User | None) -> str:
    """Obtiene nombre para mostrar del usuario (nombre completo o username)."""
    if not user:
        return ""
    nombre = getattr(user, "nombre", None) or ""
    apellido = getattr(user, "apellido", None) or ""
    full = f"{nombre} {apellido}".strip()
    return full if full else (getattr(user, "username", None) or "")


@router.get("/entries/{entry_id}", response_model=JournalEntryDetailOut)
def get_entry(entry_id: int, db: Session = Depends(get_db)):
    entry = db.query(JournalEntry).options(
        joinedload(JournalEntry.lines),
        joinedload(JournalEntry.creator),
        joinedload(JournalEntry.updater),
        joinedload(JournalEntry.poster),
        joinedload(JournalEntry.reverser),
    ).filter(JournalEntry.id == entry_id).first()
    
    if not entry:
        raise HTTPException(404, detail="Asiento no encontrado")
    
    # Obtener periodo
    period = db.query(Period).filter(Period.id == entry.period_id).first()
    
    # Construir líneas con información de cuenta
    lines_out = []
    for line in entry.lines:
        account = db.query(Account).filter(Account.id == line.account_id).first()
        lines_out.append(EntryLineOut(
            id=line.id,
            account_code=account.code if account else "",
            account_name=account.name if account else "",
            debit=float(line.debit),
            credit=float(line.credit),
            memo=line.memo,
            third_party_id=line.third_party_id,
            cost_center=line.cost_center,
        ))
    
    total_debit = float(sum([float(l.debit) for l in entry.lines]))
    total_credit = float(sum([float(l.credit) for l in entry.lines]))
    
    # Trazabilidad: fechas en ISO string
    created_at_str = entry.created_at.isoformat() if entry.created_at else None
    updated_at_str = entry.updated_at.isoformat() if entry.updated_at else None
    posted_at_str = entry.posted_at.isoformat() if entry.posted_at else None
    reversed_at_str = entry.reversed_at.isoformat() if entry.reversed_at else None
    
    return JournalEntryDetailOut(
        id=entry.id,
        company_id=entry.company_id,
        date=entry.date,
        glosa=entry.glosa or "",
        currency=entry.currency,
        origin=entry.origin,
        status=entry.status,
        total_debit=total_debit,
        total_credit=total_credit,
        period_id=entry.period_id,
        period_year=period.year if period else 0,
        period_month=period.month if period else 0,
        exchange_rate=float(entry.exchange_rate),
        lines=lines_out,
        correlative=entry.correlative,
        motor_metadata=entry.motor_metadata,
        created_by_id=entry.created_by,
        created_by_name=_user_display_name(entry.creator) if entry.creator else None,
        created_at=created_at_str,
        updated_by_id=entry.updated_by,
        updated_by_name=_user_display_name(entry.updater) if entry.updater else None,
        updated_at=updated_at_str,
        posted_by_id=entry.posted_by,
        posted_by_name=_user_display_name(entry.poster) if entry.poster else None,
        posted_at=posted_at_str,
        reversed_by_id=entry.reversed_by,
        reversed_by_name=_user_display_name(entry.reverser) if entry.reverser else None,
        reversed_at=reversed_at_str,
        integrity_hash=entry.integrity_hash,
    )

@router.post("/entries/validate", response_model=dict)
def validate_entry(
    payload: JournalEntryIn,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Valida un asiento contable usando las reglas de cuentas configuradas"""
    from ...application.services_account_rules import validate_entry_with_rules
    
    if not payload.lines or len(payload.lines) < 2:
        return {
            "is_valid": False,
            "errors": [{"message": "Un asiento debe tener al menos 2 líneas (partida doble)"}],
            "warnings": [],
            "suggestions": [],
            "compatible_accounts": []
        }
    
    # Extraer códigos de cuenta de las líneas
    account_codes = []
    for line in payload.lines:
        if line.account_code:
            account_codes.append(line.account_code)
    
    # Validar con reglas
    result = validate_entry_with_rules(db, payload.company_id, payload.lines, account_codes)
    
    return {
        "is_valid": result.is_valid,
        "errors": result.errors,
        "warnings": result.warnings,
        "suggestions": result.suggestions,
        "compatible_accounts": result.compatible_accounts
    }

@router.post("/entries", response_model=JournalEntryOut)
def post_entry(payload: JournalEntryIn, current_user: User = Depends(get_current_user)):
    if not payload.glosa or not payload.glosa.strip():
        raise HTTPException(400, detail="La glosa (descripción) es obligatoria")
    
    if not payload.lines or len(payload.lines) < 2:
        raise HTTPException(400, detail="Un asiento debe tener al menos 2 líneas (partida doble)")
    
    # Validar que el período no esté cerrado
    from ...application.services_cierre_periodo import can_modify_entry_in_period
    uow = UnitOfWork()
    
    # Obtener el período a partir de la fecha (igual que lo hace el servicio)
    y, m = payload.date.year, payload.date.month
    period = uow.periods.get_or_open(payload.company_id, y, m)
    if not period:
        uow.close()
        raise HTTPException(404, detail="Período no encontrado")
    
    role_str = str(current_user.role.value) if hasattr(current_user.role, 'value') else str(current_user.role)
    if not can_modify_entry_in_period(uow.db, period.id, role_str):
        if period.status == "CERRADO":
            uow.close()
            raise HTTPException(403, detail=f"El período {period.year}-{period.month:02d} está cerrado. No se pueden crear asientos. Un administrador debe reabrir el período primero.")
        else:
            uow.close()
            raise HTTPException(403, detail="No autorizado para crear asientos en este período")
    
    # Validar con reglas de cuentas (solo advertencias, no bloquea)
    from ...application.services_account_rules import validate_entry_with_rules
    validation_result = validate_entry_with_rules(uow.db, payload.company_id, payload.lines)
    
    # Si hay errores críticos, lanzar excepción
    if validation_result.errors:
        error_messages = [err["message"] for err in validation_result.errors]
        uow.close()
        raise HTTPException(400, detail=f"Error de validación: {'; '.join(error_messages)}")
    
    try:
        # Para asientos MANUAL, usar servicio de asientos manuales (crea DRAFT)
        # Para otros orígenes, usar servicio estándar (crea POSTED)
        origin = payload.origin or "MANUAL"
        validation_info = {}
        
        if origin == "MANUAL":
            from ...application.services_journal_manual import create_draft_entry
            entry, validation_info = create_draft_entry(
                uow=uow,
                entry_data=payload,
                user_id=current_user.id
            )
            # Combinar advertencias de ambos sistemas de validación
            all_warnings = list(validation_result.warnings) + validation_info.get("warnings", [])
        else:
            entry = __service_post(uow, payload, user_id=current_user.id)
            all_warnings = validation_result.warnings
        
        uow.commit()
        total_debit = float(sum([float(l.debit) for l in entry.lines]))
        total_credit = float(sum([float(l.credit) for l in entry.lines]))
        
        # Incluir advertencias y sugerencias en la respuesta si existen
        response_data = {
            "id": entry.id,
            "company_id": entry.company_id,
            "date": entry.date,
            "glosa": entry.glosa or "",
            "currency": entry.currency,
            "origin": entry.origin,
            "status": entry.status,
            "total_debit": total_debit,
            "total_credit": total_credit,
            "reversed_entry_id": entry.reversed_entry_id,
        }
        
        # Agregar advertencias y sugerencias si existen
        if all_warnings or validation_result.suggestions:
            response_data["validation_warnings"] = all_warnings
            response_data["validation_suggestions"] = validation_result.suggestions
            if origin == "MANUAL" and validation_info:
                response_data["requires_confirmation"] = validation_info.get("requires_confirmation", False)
        
        return JournalEntryOut(**response_data)
    finally:
        uow.close()

@router.patch("/entries/{entry_id}", response_model=JournalEntryOut)
def patch_entry(entry_id: int, payload: JournalEntryIn, current_user: User = Depends(get_current_user)):
    """
    Actualiza un asiento contable existente.
    
    REGLAS TIPO SAP:
    - Solo se pueden editar asientos en estado DRAFT
    - Los asientos POSTED son inmutables (deben revertirse)
    """
    if not payload.glosa or not payload.glosa.strip():
        raise HTTPException(400, detail="La glosa (descripción) es obligatoria")
    
    if not payload.lines or len(payload.lines) < 2:
        raise HTTPException(400, detail="Un asiento debe tener al menos 2 líneas (partida doble)")
    
    # Validar que el período no esté cerrado
    from ...application.services_cierre_periodo import can_modify_entry_in_period
    uow = UnitOfWork()
    
    # Obtener el asiento existente para verificar su período
    existing_entry = uow.db.query(JournalEntry).filter(JournalEntry.id == entry_id).first()
    if not existing_entry:
        uow.close()
        raise HTTPException(404, detail="Asiento no encontrado")
    
    # VALIDACIÓN CRÍTICA TIPO SAP: Solo DRAFT es editable
    if existing_entry.status != "DRAFT":
        uow.close()
        raise HTTPException(
            400,
            detail=f"No se puede modificar el asiento #{entry_id}. "
                   f"Estado actual: {existing_entry.status}. "
                   f"Solo se pueden editar asientos en estado DRAFT. "
                   f"Para corregir un asiento POSTED, debe crear un asiento de reversión."
        )
    
    # Obtener el período a partir de la fecha nueva (o mantener el período del asiento si no cambió la fecha)
    # El servicio post_journal_entry también obtiene el período de la fecha
    y, m = payload.date.year, payload.date.month
    period = uow.periods.get_or_open(existing_entry.company_id, y, m)
    if not period:
        uow.close()
        raise HTTPException(404, detail="Período no encontrado")
    
    role_str = str(current_user.role.value) if hasattr(current_user.role, 'value') else str(current_user.role)
    if not can_modify_entry_in_period(uow.db, period.id, role_str):
        if period.status == "CERRADO":
            uow.close()
            raise HTTPException(403, detail=f"El período {period.year}-{period.month:02d} está cerrado. No se pueden editar asientos. Un administrador debe reabrir el período primero.")
        else:
            uow.close()
            raise HTTPException(403, detail="No autorizado para editar asientos en este período")
    try:
        entry = __service_patch(uow, entry_id, payload)
        uow.commit()
        total_debit = float(sum([float(l.debit) for l in entry.lines]))
        total_credit = float(sum([float(l.credit) for l in entry.lines]))
        return JournalEntryOut(
            id=entry.id,
            company_id=entry.company_id,
            date=entry.date,
            glosa=entry.glosa or "",
            currency=entry.currency,
            origin=entry.origin,
            status=entry.status,
            total_debit=total_debit,
            total_credit=total_credit,
            reversed_entry_id=entry.reversed_entry_id,
        )
    except ValueError as e:
        raise HTTPException(400, detail=str(e))
    finally:
        uow.close()

@router.patch("/entries/{entry_id}/void", response_model=JournalEntryOut)
def void_entry(entry_id: int, db: Session = Depends(get_db), current_user: User = Depends(get_current_user)):
    entry = db.query(JournalEntry).filter(JournalEntry.id == entry_id).first()
    if not entry:
        raise HTTPException(404, detail="Asiento no encontrado")
    
    if entry.status == "VOIDED":
        raise HTTPException(400, detail="El asiento ya está anulado")
    
    # Validar que el período no esté cerrado
    from ...application.services_cierre_periodo import can_modify_entry_in_period
    period = db.query(Period).filter(Period.id == entry.period_id).first()
    role_str = str(current_user.role.value) if hasattr(current_user.role, 'value') else str(current_user.role)
    if period and not can_modify_entry_in_period(db, entry.period_id, role_str):
        if period.status == "CERRADO":
            raise HTTPException(403, detail=f"El período {period.year}-{period.month:02d} está cerrado. No se pueden anular asientos. Un administrador debe reabrir el período primero.")
        else:
            raise HTTPException(403, detail="No autorizado para anular asientos en este período")
    
    # Validar si el asiento tiene líneas conciliadas
    from ...domain.models import BankTransaction, EntryLine
    reconciled_lines = (
        db.query(BankTransaction)
        .join(EntryLine, EntryLine.id == BankTransaction.entry_line_id)
        .filter(EntryLine.entry_id == entry_id)
        .filter(BankTransaction.reconciled == True)
        .count()
    )
    
    if reconciled_lines > 0:
        raise HTTPException(400, detail=f"No se puede anular el asiento porque tiene {reconciled_lines} línea(s) conciliada(s). Primero debe revertir la(s) conciliación(es) desde el módulo de Conciliación Bancaria.")
    
    # Validar si el asiento fue generado automáticamente (no manual)
    if entry.origin != "MANUAL":
        # Verificar si está relacionado con compras
        from ...domain.models_ext import Purchase
        purchase = db.query(Purchase).filter(Purchase.journal_entry_id == entry_id).first()
        if purchase:
            raise HTTPException(400, detail=f"No se puede anular el asiento porque está relacionado con la compra #{purchase.id}. Debe eliminar o modificar la compra primero.")
        
        # Verificar si está relacionado con ventas
        from ...domain.models_ext import Sale
        sale = db.query(Sale).filter(Sale.journal_entry_id == entry_id).first()
        if sale:
            raise HTTPException(400, detail=f"No se puede anular el asiento porque está relacionado con la venta #{sale.id}. Debe eliminar o modificar la venta primero.")
        
        # Verificar si está relacionado con movimientos de inventario
        from ...domain.models_ext import InventoryMovement
        inventory_movement = db.query(InventoryMovement).filter(InventoryMovement.journal_entry_id == entry_id).first()
        if inventory_movement:
            raise HTTPException(400, detail=f"No se puede anular el asiento porque está relacionado con un movimiento de inventario. Debe revertir el movimiento primero.")
        
        # Si es automático pero no tiene relaciones específicas, permitir anular con advertencia
        # (esto permite anular asientos automáticos huérfanos)
    
    entry.status = "VOIDED"
    db.commit()
    db.refresh(entry)
    
    total_debit = float(sum([float(l.debit) for l in entry.lines]))
    total_credit = float(sum([float(l.credit) for l in entry.lines]))
    
    return JournalEntryOut(
        id=entry.id,
        company_id=entry.company_id,
        date=entry.date,
        glosa=entry.glosa or "",
        currency=entry.currency,
        origin=entry.origin,
        status=entry.status,
        total_debit=total_debit,
        total_credit=total_credit,
        reversed_entry_id=entry.reversed_entry_id,
    )

@router.patch("/entries/{entry_id}/reactivate", response_model=JournalEntryOut)
def reactivate_entry(entry_id: int, db: Session = Depends(get_db)):
    """Reactiva un asiento anulado"""
    entry = db.query(JournalEntry).filter(JournalEntry.id == entry_id).first()
    if not entry:
        raise HTTPException(404, detail="Asiento no encontrado")
    
    if entry.status != "VOIDED":
        raise HTTPException(400, detail="El asiento no está anulado")
    
    entry.status = "POSTED"
    db.commit()
    db.refresh(entry)
    
    total_debit = float(sum([float(l.debit) for l in entry.lines]))
    total_credit = float(sum([float(l.credit) for l in entry.lines]))
    
    return JournalEntryOut(
        id=entry.id,
        company_id=entry.company_id,
        date=entry.date,
        glosa=entry.glosa or "",
        currency=entry.currency,
        origin=entry.origin,
        status=entry.status,
        total_debit=total_debit,
        total_credit=total_credit,
        reversed_entry_id=entry.reversed_entry_id,
    )

def __service_post(uow: UnitOfWork, payload: JournalEntryIn, user_id: int | None = None):
    from ...application.services import post_journal_entry
    return post_journal_entry(uow, payload, user_id=user_id)

def __service_patch(uow: UnitOfWork, entry_id: int, payload: JournalEntryIn):
    from ...application.services import patch_journal_entry
    return patch_journal_entry(uow, entry_id, payload)

# ===== AI SUGGESTIONS =====

ACCOUNT_KEYWORDS = {
    "caja": ["10.1"], "efectivo": ["10.1"], "dinero": ["10.1"],
    "banco": ["10.2"], "cuenta bancaria": ["10.2"],
    "cuenta por cobrar": ["12.1"], "factura por cobrar": ["12.1"], "cobrar": ["12.1"],
    "mercaderia": ["20.1"], "inventario": ["20.1"], "stock": ["20.1"],
    "cuenta por pagar": ["42.1"], "factura por pagar": ["42.1"], "pagar": ["42.1"],
    "igv": ["40.1"], "impuesto": ["40.1"],
    "sueldo": ["41.1"], "salario": ["41.1"],
    "capital": ["50.1"],
    "venta": ["70.1"], "ingreso": ["70.1"],
    "compra": ["60.1"], "gasto": ["91"], "costo": ["90.1"],
    "servicio": ["63.1"], "alquiler": ["63.1"], "luz": ["63.1"], "agua": ["63.1"],
}

@router.get("/suggest-accounts")
def suggest_accounts(
    query: str = Query(..., description="Texto de búsqueda"),
    company_id: int = Query(..., description="ID de la empresa"),
    db: Session = Depends(get_db),
):
    """Sugiere cuentas basándose en palabras clave."""
    query_lower = query.lower()
    suggested_codes = set()
    
    for keyword, codes in ACCOUNT_KEYWORDS.items():
        if keyword in query_lower:
            suggested_codes.update(codes)
    
    accounts = db.query(Account).filter(Account.company_id == company_id, Account.active == True).all()
    for acc in accounts:
        if any(word in acc.name.lower() for word in query_lower.split() if len(word) > 3):
            suggested_codes.add(acc.code)
    
    result = []
    for code in list(suggested_codes)[:10]:
        acc = db.query(Account).filter(Account.company_id == company_id, Account.code == code).first()
        if acc:
            result.append({"code": acc.code, "name": acc.name, "type": acc.type.value})
    
    return {"suggestions": result}

@router.get("/suggest-entry")
def suggest_entry(
    glosa: str = Query(..., description="Glosa/descripción del asiento"),
    company_id: int = Query(..., description="ID de la empresa"),
    monto: float = Query(default=None, description="Monto estimado (opcional)"),
    db: Session = Depends(get_db),
):
    """
    Sugiere un asiento completo basándose en la glosa.
    Basado en PCGE y estructura contable peruana.
    """
    from ...domain.enums import AccountType
    
    glosa_lower = glosa.lower()
    suggested_lines = []
    
    # Obtener todas las cuentas activas de la empresa
    all_accounts = db.query(Account).filter(Account.company_id == company_id, Account.active == True).all()
    
    def find_account_by_keywords(keywords: List[str], account_type: str = None) -> Account | None:
        """Busca cuenta por palabras clave en código o nombre."""
        for acc in all_accounts:
            if account_type and acc.type.value != account_type:
                continue
            acc_text = (acc.code + " " + acc.name).lower()
            if any(kw.lower() in acc_text for kw in keywords):
                return acc
        # Buscar por código que empiece con el grupo (ej: 42, 12, etc.)
        if account_type:
            for kw in keywords:
                if len(kw) >= 2 and kw[:2].isdigit():
                    prefix = kw[:2]
                    for acc in all_accounts:
                        if acc.type.value == account_type and acc.code.startswith(prefix):
                            return acc
        return None
    
    # Detectar tipo de transacción según PCGE peruano
    if any(word in glosa_lower for word in ["pago", "pagamos", "pagado", "pagar"]):
        if any(word in glosa_lower for word in ["proveedor", "factura", "f001", "f-001", "f 001"]):
            # PAGO A PROVEEDOR - PCGE: 42 (Cuentas por Pagar) Debe -> 10 (Bancos/Caja) Crédito
            cuenta_pagar = find_account_by_keywords(["42", "proveedor", "pagar"], "P")
            cuenta_pago = find_account_by_keywords(["10.2", "banco"], "A") or find_account_by_keywords(["10.1", "caja"], "A") or find_account_by_keywords(["10"], "A")
            if cuenta_pagar and cuenta_pago:
                suggested_lines = [
                    {"account_code": cuenta_pagar.code, "side": "debit", "amount": monto},
                    {"account_code": cuenta_pago.code, "side": "credit", "amount": monto},
                ]
    
    elif any(word in glosa_lower for word in ["cobro", "cobramos", "cobrado", "recibido", "recibimos"]):
        # COBRO DE CLIENTE - PCGE: 10 (Bancos/Caja) Debe -> 12 (Cuentas por Cobrar) Crédito
        cuenta_cobro = find_account_by_keywords(["10.2", "banco"], "A") or find_account_by_keywords(["10.1", "caja"], "A") or find_account_by_keywords(["10"], "A")
        cuenta_cobrar = find_account_by_keywords(["12", "cobrar", "cliente"], "A")
        if cuenta_cobro and cuenta_cobrar:
            suggested_lines = [
                {"account_code": cuenta_cobro.code, "side": "debit", "amount": monto},
                {"account_code": cuenta_cobrar.code, "side": "credit", "amount": monto},
            ]
    
    elif "compra" in glosa_lower or "comprar" in glosa_lower:
        # COMPRA CON IGV - PCGE: 60.11 (Compras) + 40.11 (IGV) Debe -> 42.12 (Proveedores) Crédito
        if monto:
            base = monto / 1.18
            igv = monto - base
            cuenta_compra = find_account_by_keywords(["60", "compra"], "G")
            cuenta_igv = find_account_by_keywords(["40.1", "igv", "40"], "P")
            cuenta_proveedor = find_account_by_keywords(["42", "proveedor"], "P")
            
            if cuenta_compra and cuenta_proveedor:
                lines_to_add = [{"account_code": cuenta_compra.code, "side": "debit", "amount": base}]
                if cuenta_igv:
                    lines_to_add.append({"account_code": cuenta_igv.code, "side": "debit", "amount": igv})
                lines_to_add.append({"account_code": cuenta_proveedor.code, "side": "credit", "amount": monto})
                suggested_lines = lines_to_add
        else:
            cuenta_compra = find_account_by_keywords(["60", "compra"], "G")
            cuenta_proveedor = find_account_by_keywords(["42", "proveedor"], "P")
            if cuenta_compra and cuenta_proveedor:
                suggested_lines = [
                    {"account_code": cuenta_compra.code, "side": "debit", "amount": None},
                    {"account_code": cuenta_proveedor.code, "side": "credit", "amount": None},
                ]
    
    elif any(word in glosa_lower for word in ["venta", "vendimos", "facturar", "factura emitida", "vendemos"]):
        # VENTA CON IGV - PCGE: 12 (Clientes) Debe -> 70 (Ventas) + 40 (IGV) Crédito
        if monto:
            base = monto / 1.18
            igv = monto - base
            cuenta_cliente = find_account_by_keywords(["12", "cliente", "cobrar"], "A")
            cuenta_venta = find_account_by_keywords(["70", "venta"], "I")
            cuenta_igv = find_account_by_keywords(["40.1", "igv", "40"], "P")
            
            if cuenta_cliente and cuenta_venta:
                lines_to_add = [{"account_code": cuenta_cliente.code, "side": "debit", "amount": monto}]
                if cuenta_venta:
                    lines_to_add.append({"account_code": cuenta_venta.code, "side": "credit", "amount": base})
                if cuenta_igv:
                    lines_to_add.append({"account_code": cuenta_igv.code, "side": "credit", "amount": igv})
                suggested_lines = lines_to_add
        else:
            cuenta_cliente = find_account_by_keywords(["12", "cliente"], "A")
            cuenta_venta = find_account_by_keywords(["70", "venta"], "I")
            if cuenta_cliente and cuenta_venta:
                suggested_lines = [
                    {"account_code": cuenta_cliente.code, "side": "debit", "amount": None},
                    {"account_code": cuenta_venta.code, "side": "credit", "amount": None},
                ]
    
    # Completar información de cuentas encontradas
    valid_lines = []
    for line in suggested_lines:
        acc = db.query(Account).filter(Account.company_id == company_id, Account.code == line["account_code"]).first()
        if acc:
            line["account_name"] = acc.name
            valid_lines.append(line)
    
    return {"suggested_lines": valid_lines, "confidence": "high" if len(valid_lines) >= 2 else "low"}

@router.get("/similar-entries")
def get_similar_entries(
    glosa: str = Query(..., description="Glosa para buscar asientos similares"),
    company_id: int = Query(..., description="ID de la empresa"),
    limit: int = Query(default=5, ge=1, le=20),
    db: Session = Depends(get_db),
):
    """
    Busca asientos similares basándose en la glosa.
    Útil para reutilizar asientos frecuentes.
    """
    glosa_words = set(word.lower() for word in glosa.split() if len(word) > 3)
    
    # Buscar asientos con glosas similares
    all_entries = db.query(JournalEntry).filter(
        JournalEntry.company_id == company_id,
        JournalEntry.status == "POSTED"
    ).order_by(JournalEntry.date.desc()).limit(100).all()
    
    similar = []
    for entry in all_entries:
        entry_words = set(word.lower() for word in (entry.glosa or "").split() if len(word) > 3)
        if entry_words:
            similarity = len(glosa_words & entry_words) / max(len(glosa_words), 1)
            if similarity > 0.3:  # Al menos 30% de palabras comunes
                similar.append((similarity, entry))
    
    # Ordenar por similitud
    similar.sort(key=lambda x: x[0], reverse=True)
    
    result = []
    for similarity, entry in similar[:limit]:
        total_debit = float(sum([float(l.debit) for l in entry.lines]))
        total_credit = float(sum([float(l.credit) for l in entry.lines]))
        
        # Obtener primeras líneas como ejemplo
        sample_lines = []
        for line in entry.lines[:3]:
            acc = db.query(Account).filter(Account.id == line.account_id).first()
            if acc:
                sample_lines.append({
                    "account_code": acc.code,
                    "account_name": acc.name,
                    "debit": float(line.debit),
                    "credit": float(line.credit),
                })
        
        result.append({
            "id": entry.id,
            "date": entry.date.isoformat(),
            "glosa": entry.glosa,
            "similarity": round(similarity * 100, 1),
            "total_debit": total_debit,
            "total_credit": total_credit,
            "sample_lines": sample_lines,
        })
    
    return {"similar_entries": result}

@router.get("/templates")
def get_templates(company_id: int = Query(..., description="ID de la empresa"), db: Session = Depends(get_db)):
    """
    Retorna plantillas predefinidas de asientos comunes según PCGE peruano.
    Útil para usuarios sin experiencia en contabilidad.
    """
    from ...domain.enums import AccountType
    
    all_accounts = db.query(Account).filter(Account.company_id == company_id, Account.active == True).all()
    
    def find_account(code_prefix: str, keywords: List[str], acc_type: str = None) -> Account | None:
        for acc in all_accounts:
            if acc_type and acc.type.value != acc_type:
                continue
            acc_text = (acc.code + " " + acc.name).lower()
            if code_prefix and acc.code.startswith(code_prefix):
                return acc
            if any(kw.lower() in acc_text for kw in keywords):
                return acc
        return None
    
    templates = []
    
    # Plantilla 1: Pago a Proveedor
    cuenta_pagar = find_account("42", ["proveedor", "pagar"], "P")
    cuenta_pago = find_account("10.2", ["banco"], "A") or find_account("10.1", ["caja"], "A") or find_account("10", [], "A")
    if cuenta_pagar and cuenta_pago:
        templates.append({
            "id": "pago_proveedor",
            "name": "Pago a Proveedor",
            "description": "Registra el pago de una factura a proveedor",
            "glosa_example": "Pago a proveedor ABC por factura F001-0001",
            "lines": [
                {"account_code": cuenta_pagar.code, "account_name": cuenta_pagar.name, "side": "debit", "description": "Reducimos nuestra deuda con el proveedor"},
                {"account_code": cuenta_pago.code, "account_name": cuenta_pago.name, "side": "credit", "description": "Salida de dinero del banco/caja"},
            ]
        })
    
    # Plantilla 2: Cobro de Cliente
    cuenta_cobro = find_account("10.2", ["banco"], "A") or find_account("10.1", ["caja"], "A") or find_account("10", [], "A")
    cuenta_cobrar = find_account("12", ["cobrar", "cliente"], "A")
    if cuenta_cobro and cuenta_cobrar:
        templates.append({
            "id": "cobro_cliente",
            "name": "Cobro de Cliente",
            "description": "Registra el cobro de una factura a cliente",
            "glosa_example": "Cobro de cliente XYZ por factura F001-0002",
            "lines": [
                {"account_code": cuenta_cobro.code, "account_name": cuenta_cobro.name, "side": "debit", "description": "Entrada de dinero al banco/caja"},
                {"account_code": cuenta_cobrar.code, "account_name": cuenta_cobrar.name, "side": "credit", "description": "Reducimos lo que nos deben los clientes"},
            ]
        })
    
    # Plantilla 3: Compra con IGV
    cuenta_compra = find_account("60", ["compra"], "G")
    cuenta_igv = find_account("40.1", ["igv"], "P") or find_account("40", [], "P")
    cuenta_proveedor = find_account("42", ["proveedor"], "P")
    if cuenta_compra and cuenta_proveedor:
        templates.append({
            "id": "compra_igv",
            "name": "Compra con IGV",
            "description": "Registra una compra con IGV incluido (automático: base + IGV)",
            "glosa_example": "Compra de mercadería a proveedor ABC",
            "lines": [
                {"account_code": cuenta_compra.code, "account_name": cuenta_compra.name, "side": "debit", "description": "Costo de lo comprado (sin IGV)", "auto_calculate": "base"},
                {"account_code": cuenta_igv.code if cuenta_igv else "", "account_name": cuenta_igv.name if cuenta_igv else "IGV", "side": "debit", "description": "IGV crédito fiscal (18%)", "auto_calculate": "igv", "optional": True},
                {"account_code": cuenta_proveedor.code, "account_name": cuenta_proveedor.name, "side": "credit", "description": "Deuda con el proveedor (total con IGV)", "auto_calculate": "total"},
            ]
        })
    
    # Plantilla 4: Venta con IGV
    cuenta_cliente = find_account("12", ["cliente", "cobrar"], "A")
    cuenta_venta = find_account("70", ["venta"], "I")
    if cuenta_cliente and cuenta_venta:
        templates.append({
            "id": "venta_igv",
            "name": "Venta con IGV",
            "description": "Registra una venta con IGV incluido (automático: base + IGV)",
            "glosa_example": "Venta de productos a cliente XYZ",
            "lines": [
                {"account_code": cuenta_cliente.code, "account_name": cuenta_cliente.name, "side": "debit", "description": "Lo que nos deben los clientes (total con IGV)", "auto_calculate": "total"},
                {"account_code": cuenta_venta.code, "account_name": cuenta_venta.name, "side": "credit", "description": "Ingreso por ventas (sin IGV)", "auto_calculate": "base"},
                {"account_code": cuenta_igv.code if cuenta_igv else "", "account_name": cuenta_igv.name if cuenta_igv else "IGV", "side": "credit", "description": "IGV débito fiscal (18%)", "auto_calculate": "igv", "optional": True},
            ]
        })
    
    # Plantilla 5: Pago de Servicios
    cuenta_servicio = find_account("63", ["servicio"], "G") or find_account("91", ["gasto"], "G")
    cuenta_pago_serv = find_account("10.2", ["banco"], "A") or find_account("10.1", ["caja"], "A")
    if cuenta_servicio and cuenta_pago_serv:
        templates.append({
            "id": "pago_servicio",
            "name": "Pago de Servicios",
            "description": "Registra el pago de servicios (luz, agua, telefonía, etc.)",
            "glosa_example": "Pago de servicio de luz - mes octubre",
            "lines": [
                {"account_code": cuenta_servicio.code, "account_name": cuenta_servicio.name, "side": "debit", "description": "Gasto por servicio"},
                {"account_code": cuenta_pago_serv.code, "account_name": cuenta_pago_serv.name, "side": "credit", "description": "Salida de dinero"},
            ]
        })
    
    return {"templates": templates}

@router.get("/entries/export/excel")
def export_entries_excel(
    company_id: int = Query(..., description="ID de la empresa"),
    period_id: int | None = Query(default=None),
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    account_code: str | None = Query(default=None),
    status: str | None = Query(default=None),
    db: Session = Depends(get_db),
):
    """Exporta el libro diario a Excel - Los datos se consultan directamente desde la BD, no desde el grid"""
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        logger.info(f"Exportando Excel - company_id={company_id}, period_id={period_id}, account_code={account_code}")
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        
        # Obtener entradas (misma lógica que list_entries) con relaciones cargadas
        # Primero filtrar por account_code si es necesario para obtener los IDs
        entry_ids = None
        if account_code:
            subquery = db.query(JournalEntry.id).join(EntryLine).join(Account).filter(
                JournalEntry.company_id == company_id,
                Account.code == account_code
            ).distinct()
            entry_ids = [row[0] for row in subquery.all()]
            if not entry_ids:
                raise HTTPException(status_code=404, detail="No hay asientos para exportar con los filtros seleccionados")
        
        # Construir query principal con joinedload
        query = db.query(JournalEntry).options(
            joinedload(JournalEntry.lines)
        ).filter(JournalEntry.company_id == company_id)
        
        if entry_ids:
            query = query.filter(JournalEntry.id.in_(entry_ids))
        if period_id:
            query = query.filter(JournalEntry.period_id == period_id)
        if date_from:
            query = query.filter(JournalEntry.date >= date_from)
        if date_to:
            query = query.filter(JournalEntry.date <= date_to)
        if status:
            query = query.filter(JournalEntry.status == status)
        
        entries = query.order_by(JournalEntry.date, JournalEntry.id).all()
        
        if not entries:
            raise HTTPException(status_code=404, detail="No hay asientos para exportar con los filtros seleccionados")
        
        # Verificar que hay datos para exportar
        has_data = False
        for entry in entries:
            if entry.lines:
                has_data = True
                break
        
        if not has_data:
            raise HTTPException(status_code=404, detail="No hay datos para exportar. Los asientos no tienen líneas.")
        
        # Pre-cargar todas las cuentas necesarias en una sola query
        account_ids = set()
        for entry in entries:
            if entry.lines:
                for line in entry.lines:
                    if line.account_id:
                        account_ids.add(line.account_id)
        
        accounts_dict = {}
        if account_ids:
            accounts = db.query(Account).filter(Account.id.in_(list(account_ids))).all()
            accounts_dict = {acc.id: acc for acc in accounts}
        
        # Obtener información de la empresa
        from ...domain.models import Company
        company = db.query(Company).filter(Company.id == company_id).first()
        if not company:
            raise HTTPException(status_code=404, detail="Empresa no encontrada")
        
        # Obtener información del período
        period_info = ""
        if period_id:
            period = db.query(Period).filter(Period.id == period_id).first()
            if period:
                period_info = f"{period.year}-{period.month:02d}"
        elif date_from and date_to:
            period_info = f"{date_from} al {date_to}"
        
        # Crear workbook
        wb = Workbook()
        ws = wb.active
        ws.title = "Libro Diario"
        
        # Encabezado formato SUNAT
        title_font = Font(bold=True, size=14)
        header_font = Font(bold=True, size=11)
        
        ws.append(["LIBRO DIARIO - FORMATO SUNAT"])
        ws.merge_cells('A1:G1')
        ws['A1'].font = title_font
        ws['A1'].alignment = Alignment(horizontal="center")
        ws.append([])
        
        ws.append(["Empresa:", company.name])
        ws.merge_cells('A3:G3')
        ws['A3'].font = header_font
        
        if company.ruc:
            ws.append(["RUC:", company.ruc])
            ws.merge_cells('A4:G4')
            ws['A4'].font = Font(size=11)
        
        ws.append(["Período:", period_info])
        ws.merge_cells('A5:G5')
        ws['A5'].font = Font(size=11)
        
        ws.append([])
        
        # Headers de columnas
        headers = ["Nro.", "Fecha", "Glosa", "Cuenta", "Nombre Cuenta", "Debe", "Haber"]
        ws.append(headers)
        
        # Estilo de headers (la fila de headers depende de si hay RUC)
        header_row = 7 if company.ruc else 6
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        for cell in ws[header_row]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Datos con formato SUNAT
        row_count = 0
        correlativo = 1
        total_debe = 0.0
        total_haber = 0.0
        
        for entry in entries:
            if not entry.lines or entry.status == "VOIDED":
                continue  # Saltar entradas sin líneas o anuladas
            
            for line in entry.lines:
                acc = accounts_dict.get(line.account_id) if line.account_id else None
                if not acc:
                    continue
                
                debit_val = float(line.debit) if line.debit else 0.0
                credit_val = float(line.credit) if line.credit else 0.0
                total_debe += debit_val
                total_haber += credit_val
                
                glosa_text = (entry.glosa or "")[:100]  # Limitar glosa a 100 caracteres
                
                ws.append([
                    correlativo,
                    entry.date.strftime('%d/%m/%Y') if entry.date else "",
                    glosa_text,
                    acc.code,
                    acc.name[:50],  # Limitar nombre a 50 caracteres
                    debit_val,
                    credit_val,
                ])
                correlativo += 1
                row_count += 1
        
        # Fila de totales
        ws.append([])
        ws.append(["TOTALES", "", "", "", "", total_debe, total_haber])
        
        # Estilo de totales
        total_row = ws.max_row
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G']:
            cell = ws[f'{col}{total_row}']
            if col == 'A':
                cell.font = Font(bold=True, size=11)
            elif col in ['F', 'G']:
                cell.font = Font(bold=True, size=11)
                cell.number_format = '#,##0.00'
            cell.fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        # Auto-ajustar columnas
        column_widths = {'A': 12, 'B': 12, 'C': 50, 'D': 15, 'E': 40, 'F': 15, 'G': 15}
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # Formato numérico para columnas Debe y Haber
        data_start_row = header_row + 1
        for row in range(data_start_row, ws.max_row):  # Desde la fila de datos hasta antes de totales
            ws[f'F{row}'].number_format = '#,##0.00'
            ws[f'G{row}'].number_format = '#,##0.00'
        
        # Guardar a BytesIO
        buffer = BytesIO()
        try:
            wb.save(buffer)
            # IMPORTANTE: flush y obtener posición antes de leer
            buffer.flush()
            buffer.seek(0)  # Ir al inicio
            buffer_content = buffer.read()  # Leer todo el contenido
            buffer.close()  # Cerrar el buffer
        except Exception as e:
            logger.error(f"Error al guardar Excel al buffer: {str(e)}")
            raise HTTPException(status_code=500, detail=f"Error al guardar Excel: {str(e)}")
        
        # Verificar que el buffer tiene contenido
        buffer_size = len(buffer_content)
        logger.info(f"Excel generado - tamaño: {buffer_size} bytes, filas: {row_count}")
        
        if buffer_size == 0:
            logger.error(f"Excel vacío - row_count={row_count}")
            raise HTTPException(status_code=500, detail=f"Error: El archivo Excel generado está vacío. Se agregaron {row_count} filas de datos.")
        
        if buffer_size < 500:
            logger.warning(f"Excel muy pequeño - {buffer_size} bytes")
        
        return Response(
            content=buffer_content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="libro_diario_{date.today().isoformat()}.xlsx"',
                "Content-Length": str(buffer_size)
            }
        )
    except Exception as e:
        import traceback
        logger.error(f"Error al exportar a Excel: {str(e)}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error al exportar a Excel: {str(e)}")

@router.get("/entries/export/pdf")
def export_entries_pdf(
    company_id: int = Query(..., description="ID de la empresa"),
    period_id: int | None = Query(default=None),
    date_from: date | None = Query(default=None),
    date_to: date | None = Query(default=None),
    account_code: str | None = Query(default=None),
    status: str | None = Query(default=None),
    db: Session = Depends(get_db),
):
    """Exporta el libro diario a PDF - Los datos se consultan directamente desde la BD, no desde el grid"""
    import logging
    logger = logging.getLogger(__name__)
    
    try:
        logger.info(f"Exportando PDF - company_id={company_id}, period_id={period_id}, account_code={account_code}")
        from reportlab.lib import colors
        from reportlab.lib.pagesizes import letter, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet
        from reportlab.lib.units import inch
        
        # Obtener entradas con relaciones cargadas
        # Primero filtrar por account_code si es necesario para obtener los IDs
        entry_ids = None
        if account_code:
            subquery = db.query(JournalEntry.id).join(EntryLine).join(Account).filter(
                JournalEntry.company_id == company_id,
                Account.code == account_code
            ).distinct()
            entry_ids = [row[0] for row in subquery.all()]
            if not entry_ids:
                raise HTTPException(status_code=404, detail="No hay asientos para exportar con los filtros seleccionados")
        
        # Construir query principal con joinedload
        query = db.query(JournalEntry).options(
            joinedload(JournalEntry.lines)
        ).filter(JournalEntry.company_id == company_id)
        
        if entry_ids:
            query = query.filter(JournalEntry.id.in_(entry_ids))
        if period_id:
            query = query.filter(JournalEntry.period_id == period_id)
        if date_from:
            query = query.filter(JournalEntry.date >= date_from)
        if date_to:
            query = query.filter(JournalEntry.date <= date_to)
        if status:
            query = query.filter(JournalEntry.status == status)
        
        entries = query.order_by(JournalEntry.date, JournalEntry.id).all()
        
        logger.info(f"Encontrados {len(entries)} asientos para exportar")
        
        if not entries:
            raise HTTPException(status_code=404, detail="No hay asientos para exportar con los filtros seleccionados")
        
        # Pre-cargar todas las cuentas necesarias en una sola query
        account_ids = set()
        for entry in entries:
            if entry.lines:
                for line in entry.lines:
                    if line.account_id:
                        account_ids.add(line.account_id)
        
        accounts_dict = {}
        if account_ids:
            accounts = db.query(Account).filter(Account.id.in_(list(account_ids))).all()
            accounts_dict = {acc.id: acc for acc in accounts}
        
        # Verificar que hay datos para exportar
        has_data = False
        for entry in entries:
            if entry.lines:
                has_data = True
                break
        
        if not has_data:
            raise HTTPException(status_code=404, detail="No hay datos para exportar. Los asientos no tienen líneas.")
        
        # Obtener información de la empresa
        from ...domain.models import Company
        company = db.query(Company).filter(Company.id == company_id).first()
        if not company:
            raise HTTPException(status_code=404, detail="Empresa no encontrada")
        
        # Preparar datos con formato SUNAT
        table_rows = []
        correlativo = 1
        total_debe = 0.0
        total_haber = 0.0
        
        for entry in entries:
            if not entry.lines or entry.status == "VOIDED":
                continue  # Saltar entradas sin líneas o anuladas
            
            for line in entry.lines:
                acc = accounts_dict.get(line.account_id) if line.account_id else None
                if not acc:
                    # Saltar líneas sin cuenta válida
                    continue
                debit_val = float(line.debit) if line.debit else 0.0
                credit_val = float(line.credit) if line.credit else 0.0
                total_debe += debit_val
                total_haber += credit_val
                glosa_text = (entry.glosa or "")[:100] if entry.glosa else ""
                table_rows.append([
                    str(correlativo),
                    entry.date.strftime('%d/%m/%Y') if entry.date else "",
                    glosa_text or "",
                    acc.code,
                    acc.name[:50] if acc.name else "",  # Limitar nombre a 50 caracteres
                    f"{debit_val:,.2f}",  # Formato con comas de miles
                    f"{credit_val:,.2f}",  # Formato con comas de miles
                ])
                correlativo += 1
        
        # Agregar fila de totales con formato de miles
        if table_rows:
            table_rows.append([
                "TOTALES", "", "", "", "", f"{total_debe:,.2f}", f"{total_haber:,.2f}"
            ])
        
        logger.info(f"Preparadas {len(table_rows)} filas para PDF")
        
        if not table_rows:
            total_entries = len(entries)
            total_lines = sum(len(e.lines) for e in entries if e.lines)
            logger.warning(f"No hay líneas válidas para PDF - {total_entries} asientos, {total_lines} líneas totales")
            raise HTTPException(
                status_code=404, 
                detail=f"No hay datos para exportar. Se encontraron {total_entries} asientos con {total_lines} líneas, pero ninguna línea tiene una cuenta válida asociada."
            )
        
        # Determinar período para el subtítulo
        period_str = None
        if period_id:
            period = db.query(Period).filter(Period.id == period_id).first()
            if period:
                period_str = f"{period.year}-{period.month:02d}"
        elif entries:
            first_entry = entries[0]
            if first_entry.date:
                period_str = f"{first_entry.date.year}-{first_entry.date.month:02d}"
        
        # Validar que hay datos antes de crear el PDF
        if len(table_rows) == 0:
            raise HTTPException(status_code=404, detail=f"No hay líneas de asientos para exportar. Se encontraron {len(entries)} asientos pero ninguno tiene líneas válidas con cuentas asociadas.")
        
        # Crear PDF con formato SUNAT (libro electrónico)
        from ...infrastructure.pdf_utils import create_sunat_pdf
        from reportlab.lib.units import inch
        
        try:
            buffer = create_sunat_pdf(
                company_name=company.name,
                company_ruc=company.ruc,
                report_title="LIBRO DIARIO",
                headers=["Nro.", "Fecha", "Glosa", "Cuenta", "Nombre Cuenta", "Debe", "Haber"],
                rows=table_rows,
                period=period_str,
                footer_text="Sistema Contable SISCONT"
            )
        except ValueError as ve:
            raise HTTPException(status_code=500, detail=f"Error al generar PDF: {str(ve)}")
        except Exception as e:
            import traceback
            error_trace = traceback.format_exc()
            raise HTTPException(status_code=500, detail=f"Error inesperado al generar PDF: {str(e)}. Trace: {error_trace[:500]}")
        
        # Verificar que el buffer tiene contenido
        buffer.seek(0)
        buffer_content = buffer.read()
        buffer.seek(0)
        
        if len(buffer_content) == 0:
            raise HTTPException(status_code=500, detail=f"Error: El archivo PDF generado está vacío. No se pudieron generar los datos. Hay {len(table_rows)} filas preparadas.")
        
        buffer_size = len(buffer_content)
        logger.info(f"PDF generado - tamaño: {buffer_size} bytes, filas: {len(table_rows)}")
        
        if buffer_size < 1000:  # Un PDF mínimo debería tener al menos 1KB
            logger.warning(f"PDF muy pequeño - {buffer_size} bytes")
            raise HTTPException(status_code=500, detail=f"Error: El archivo PDF generado parece estar incompleto (solo {buffer_size} bytes). Hay {len(table_rows)} filas de datos.")
        
        return Response(
            content=buffer_content,
            media_type="application/pdf",
            headers={
                "Content-Disposition": f'attachment; filename="libro_diario_{date.today().isoformat()}.pdf"',
                "Content-Length": str(buffer_size)
            }
        )
    except ImportError:
        raise HTTPException(status_code=500, detail="reportlab no está instalado. Instala con: pip install reportlab")
    except Exception as e:
        import traceback
        logger.error(f"Error al exportar a PDF: {str(e)}\n{traceback.format_exc()}")
        raise HTTPException(status_code=500, detail=f"Error al exportar a PDF: {str(e)}")
