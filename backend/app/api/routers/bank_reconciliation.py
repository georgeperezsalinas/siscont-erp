"""
Router para conciliación bancaria
"""
from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Query
from fastapi.responses import Response
from sqlalchemy.orm import Session
from typing import List, Optional
from pydantic import BaseModel
from datetime import date, datetime, timedelta
from decimal import Decimal
import random

from ...dependencies import get_db
from ...domain.models import BankAccount, BankStatement, BankTransaction, BankReconciliation, Account, Period, User, JournalEntry, EntryLine
from ...security.auth import get_current_user
from ...domain.enums import UserRole

router = APIRouter(prefix="/bank-reconciliation", tags=["bank-reconciliation"])

# ===== DTOs =====

class BankAccountIn(BaseModel):
    company_id: int
    account_id: int  # ID de la cuenta contable (10.x)
    bank_name: str
    account_number: str
    currency: str = "PEN"

class BankAccountOut(BaseModel):
    id: int
    company_id: int
    account_id: int
    bank_name: str
    account_number: str
    currency: str
    active: bool
    account_code: Optional[str] = None
    account_name: Optional[str] = None
    
    class Config:
        from_attributes = True

class BankTransactionIn(BaseModel):
    transaction_date: date
    description: str
    reference: Optional[str] = None
    debit: Decimal
    credit: Decimal
    balance: Decimal

class BankStatementIn(BaseModel):
    bank_account_id: int
    period_id: int
    statement_date: date
    opening_balance: Decimal
    closing_balance: Decimal
    transactions: List[BankTransactionIn]

class ReconciliationSummary(BaseModel):
    book_balance: Decimal  # Saldo según contabilidad
    bank_balance: Decimal  # Saldo según banco
    pending_debits: Decimal  # Cheques pendientes
    pending_credits: Decimal  # Depósitos en tránsito
    reconciled_balance: Decimal  # Saldo conciliado

class BankTransactionOut(BaseModel):
    id: int
    transaction_date: date
    description: str
    reference: Optional[str] = None
    debit: Decimal
    credit: Decimal
    balance: Decimal
    reconciled: bool
    entry_line_id: Optional[int] = None
    matched_entry_glosa: Optional[str] = None
    matched_entry_date: Optional[date] = None
    
    class Config:
        from_attributes = True

class EntryLineOut(BaseModel):
    id: int
    entry_id: int
    account_id: int
    account_code: Optional[str] = None
    account_name: Optional[str] = None
    debit: Decimal
    credit: Decimal
    memo: Optional[str] = None
    entry_date: date
    entry_glosa: str
    entry_number: Optional[str] = None
    reconciled: bool = False
    
    class Config:
        from_attributes = True

class MatchSuggestion(BaseModel):
    bank_transaction_id: int
    entry_line_id: int
    confidence: float  # 0.0 a 1.0
    reason: str  # Razón del match (monto, fecha, descripción)

class MatchRequest(BaseModel):
    bank_transaction_id: int
    entry_line_id: int

class BulkMatchRequest(BaseModel):
    matches: List[MatchRequest]

class FinalizeReconciliationRequest(BaseModel):
    pending_debits: Decimal = Decimal('0')
    pending_credits: Decimal = Decimal('0')
    notes: Optional[str] = None

# ===== Endpoints =====

@router.get("/bank-accounts", response_model=List[BankAccountOut])
def list_bank_accounts(
    company_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Lista todas las cuentas bancarias de una empresa"""
    accounts = db.query(BankAccount).filter(
        BankAccount.company_id == company_id,
        BankAccount.active == True
    ).all()
    
    result = []
    for acc in accounts:
        account = db.query(Account).filter(Account.id == acc.account_id).first()
        result.append(BankAccountOut(
            id=acc.id,
            company_id=acc.company_id,
            account_id=acc.account_id,
            bank_name=acc.bank_name,
            account_number=acc.account_number,
            currency=acc.currency,
            active=acc.active,
            account_code=account.code if account else None,
            account_name=account.name if account else None
        ))
    return result

@router.post("/bank-accounts", response_model=BankAccountOut)
def create_bank_account(
    payload: BankAccountIn,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Crea una nueva cuenta bancaria"""
    if current_user.role not in (UserRole.ADMINISTRADOR, UserRole.CONTADOR):
        raise HTTPException(403, "No autorizado")
    
    # Verificar que la cuenta contable existe y es de tipo banco (10.x)
    account = db.query(Account).filter(
        Account.id == payload.account_id,
        Account.company_id == payload.company_id,
        Account.code.like('10%')
    ).first()
    
    if not account:
        raise HTTPException(400, "La cuenta contable debe ser de tipo banco (código 10.x)")
    
    bank_account = BankAccount(
        company_id=payload.company_id,
        account_id=payload.account_id,
        bank_name=payload.bank_name,
        account_number=payload.account_number,
        currency=payload.currency
    )
    db.add(bank_account)
    db.commit()
    db.refresh(bank_account)
    
    return BankAccountOut(
        id=bank_account.id,
        company_id=bank_account.company_id,
        account_id=bank_account.account_id,
        bank_name=bank_account.bank_name,
        account_number=bank_account.account_number,
        currency=bank_account.currency,
        active=bank_account.active,
        account_code=account.code,
        account_name=account.name
    )

@router.post("/upload-statement", response_model=dict)
async def upload_statement(
    payload: BankStatementIn,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Carga un extracto bancario"""
    if current_user.role not in (UserRole.ADMINISTRADOR, UserRole.CONTADOR):
        raise HTTPException(403, "No autorizado")
    
    # Verificar cuenta bancaria
    bank_account = db.query(BankAccount).filter(BankAccount.id == payload.bank_account_id).first()
    if not bank_account:
        raise HTTPException(404, "Cuenta bancaria no encontrada")
    
    # Verificar período
    period = db.query(Period).filter(Period.id == payload.period_id).first()
    if not period:
        raise HTTPException(404, "Período no encontrado")
    
    # Crear statement
    statement = BankStatement(
        bank_account_id=payload.bank_account_id,
        period_id=payload.period_id,
        statement_date=payload.statement_date,
        opening_balance=payload.opening_balance,
        closing_balance=payload.closing_balance,
        uploaded_by=current_user.id,
        status="PENDIENTE"
    )
    db.add(statement)
    db.flush()
    
    # Agregar transacciones
    for tx in payload.transactions:
        bank_tx = BankTransaction(
            statement_id=statement.id,
            transaction_date=tx.transaction_date,
            description=tx.description,
            reference=tx.reference,
            debit=tx.debit,
            credit=tx.credit,
            balance=tx.balance,
            reconciled=False
        )
        db.add(bank_tx)
    
    db.commit()
    db.refresh(statement)
    
    return {
        "id": statement.id,
        "status": statement.status,
        "transaction_count": len(payload.transactions)
    }

@router.get("/reconciliation-summary/{bank_account_id}")
def get_reconciliation_summary(
    bank_account_id: int,
    period_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Obtiene el resumen de conciliación para una cuenta bancaria y período"""
    bank_account = db.query(BankAccount).filter(BankAccount.id == bank_account_id).first()
    if not bank_account:
        raise HTTPException(404, "Cuenta bancaria no encontrada")
    
    # Calcular saldo contable (acumulado hasta el período)
    from sqlalchemy import func
    
    period = db.query(Period).filter(Period.id == period_id).first()
    if not period:
        raise HTTPException(404, "Período no encontrado")
    
    # Obtener todos los períodos hasta el actual
    all_periods = db.query(Period).filter(
        Period.company_id == period.company_id
    ).order_by(Period.year, Period.month).all()
    
    period_ids_list = [
        p.id for p in all_periods
        if (p.year < period.year) or (p.year == period.year and p.month <= period.month)
    ]
    
    # Saldo contable = debe - haber para cuenta bancaria
    book_balance_query = (
        db.query(func.coalesce(func.sum(EntryLine.debit) - func.sum(EntryLine.credit), Decimal('0')))
        .join(JournalEntry, JournalEntry.id == EntryLine.entry_id)
        .filter(EntryLine.account_id == bank_account.account_id)
        .filter(JournalEntry.period_id.in_(period_ids_list))
        .filter(JournalEntry.status == "POSTED")
    )
    book_balance = float(book_balance_query.scalar() or Decimal('0'))
    
    # Obtener extracto bancario más reciente
    statement = db.query(BankStatement).filter(
        BankStatement.bank_account_id == bank_account_id,
        BankStatement.period_id == period_id
    ).order_by(BankStatement.statement_date.desc()).first()
    
    bank_balance = float(statement.closing_balance) if statement else 0.0
    
    # Obtener conciliación existente o crear nueva
    reconciliation = db.query(BankReconciliation).filter(
        BankReconciliation.bank_account_id == bank_account_id,
        BankReconciliation.period_id == period_id
    ).first()
    
    if reconciliation:
        pending_debits = float(reconciliation.pending_debits)
        pending_credits = float(reconciliation.pending_credits)
        reconciled_balance = float(reconciliation.reconciled_balance)
    else:
        pending_debits = 0.0
        pending_credits = 0.0
        reconciled_balance = book_balance  # Inicialmente igual al saldo contable
    
    return ReconciliationSummary(
        book_balance=Decimal(str(book_balance)),
        bank_balance=Decimal(str(bank_balance)),
        pending_debits=Decimal(str(pending_debits)),
        pending_credits=Decimal(str(pending_credits)),
        reconciled_balance=Decimal(str(reconciled_balance))
    )

@router.get("/transactions/{bank_account_id}")
def get_unreconciled_transactions(
    bank_account_id: int,
    period_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Obtiene transacciones bancarias no conciliadas para un período"""
    bank_account = db.query(BankAccount).filter(BankAccount.id == bank_account_id).first()
    if not bank_account:
        raise HTTPException(404, "Cuenta bancaria no encontrada")
    
    # Obtener extracto del período
    statement = db.query(BankStatement).filter(
        BankStatement.bank_account_id == bank_account_id,
        BankStatement.period_id == period_id
    ).order_by(BankStatement.statement_date.desc()).first()
    
    if not statement:
        return []
    
    # Obtener transacciones no conciliadas
    transactions = db.query(BankTransaction).filter(
        BankTransaction.statement_id == statement.id,
        BankTransaction.reconciled == False
    ).order_by(BankTransaction.transaction_date, BankTransaction.id).all()
    
    result = []
    for tx in transactions:
        matched_entry_glosa = None
        matched_entry_date = None
        if tx.entry_line_id:
            entry_line = db.query(EntryLine).filter(EntryLine.id == tx.entry_line_id).first()
            if entry_line:
                entry = db.query(JournalEntry).filter(JournalEntry.id == entry_line.entry_id).first()
                if entry:
                    matched_entry_glosa = entry.glosa
                    matched_entry_date = entry.date
        
        result.append(BankTransactionOut(
            id=tx.id,
            transaction_date=tx.transaction_date,
            description=tx.description,
            reference=tx.reference,
            debit=tx.debit,
            credit=tx.credit,
            balance=tx.balance,
            reconciled=tx.reconciled,
            entry_line_id=tx.entry_line_id,
            matched_entry_glosa=matched_entry_glosa,
            matched_entry_date=matched_entry_date
        ))
    
    return result

@router.get("/entry-lines/{bank_account_id}")
def get_unreconciled_entry_lines(
    bank_account_id: int,
    period_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Obtiene líneas contables pendientes de conciliación para una cuenta bancaria"""
    bank_account = db.query(BankAccount).filter(BankAccount.id == bank_account_id).first()
    if not bank_account:
        raise HTTPException(404, "Cuenta bancaria no encontrada")
    
    period = db.query(Period).filter(Period.id == period_id).first()
    if not period:
        raise HTTPException(404, "Período no encontrado")
    
    # Obtener todas las líneas contables de la cuenta bancaria en el período
    # que no estén conciliadas (no tienen BankTransaction asociado)
    entry_lines = (
        db.query(EntryLine)
        .join(JournalEntry, JournalEntry.id == EntryLine.entry_id)
        .join(Account, Account.id == EntryLine.account_id)
        .filter(EntryLine.account_id == bank_account.account_id)
        .filter(JournalEntry.period_id == period_id)
        .filter(JournalEntry.status == "POSTED")
        .order_by(JournalEntry.date, EntryLine.id)
        .all()
    )
    
    # Obtener IDs de líneas ya conciliadas
    reconciled_line_ids = set(
        db.query(BankTransaction.entry_line_id)
        .join(BankStatement, BankStatement.id == BankTransaction.statement_id)
        .filter(BankStatement.bank_account_id == bank_account_id)
        .filter(BankStatement.period_id == period_id)
        .filter(BankTransaction.entry_line_id.isnot(None))
        .filter(BankTransaction.reconciled == True)
        .distinct()
        .all()
    )
    reconciled_line_ids = {x[0] for x in reconciled_line_ids if x[0] is not None}
    
    result = []
    for line in entry_lines:
        if line.id in reconciled_line_ids:
            continue
        
        entry = db.query(JournalEntry).filter(JournalEntry.id == line.entry_id).first()
        account = db.query(Account).filter(Account.id == line.account_id).first()
        
        result.append(EntryLineOut(
            id=line.id,
            entry_id=line.entry_id,
            account_id=line.account_id,
            account_code=account.code if account else None,
            account_name=account.name if account else None,
            debit=line.debit,
            credit=line.credit,
            memo=line.memo,
            entry_date=entry.date if entry else date.today(),
            entry_glosa=entry.glosa if entry else "",
            entry_number=None,  # TODO: agregar número de asiento si existe
            reconciled=False
        ))
    
    return result

@router.get("/auto-match/{bank_account_id}")
def get_auto_match_suggestions(
    bank_account_id: int,
    period_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Genera sugerencias automáticas de matching entre transacciones bancarias y líneas contables"""
    
    # Obtener transacciones no conciliadas
    transactions = get_unreconciled_transactions(bank_account_id, period_id, db, current_user)
    
    # Obtener líneas contables no conciliadas
    entry_lines = get_unreconciled_entry_lines(bank_account_id, period_id, db, current_user)
    
    suggestions = []
    
    for tx in transactions:
        if tx.reconciled or tx.entry_line_id:
            continue
        
        tx_amount = float(tx.debit) if tx.debit > 0 else float(tx.credit)
        tx_date = tx.transaction_date
        
        best_match = None
        best_confidence = 0.0
        best_reason = ""
        
        for line in entry_lines:
            if line.reconciled:
                continue
            
            line_amount = float(line.debit) if line.debit > 0 else float(line.credit)
            line_date = line.entry_date
            
            # Calcular confianza basada en monto y fecha
            confidence = 0.0
            reasons = []
            
            # Match por monto exacto (peso: 0.7)
            if abs(tx_amount - line_amount) < 0.01:
                confidence += 0.7
                reasons.append("monto exacto")
            
            # Match por monto cercano (peso: 0.3)
            elif abs(tx_amount - line_amount) / max(abs(tx_amount), 0.01) < 0.05:  # 5% de diferencia
                confidence += 0.3
                reasons.append("monto cercano")
            
            # Match por fecha cercana (peso: 0.2)
            date_diff = abs((tx_date - line_date).days)
            if date_diff <= 3:
                confidence += 0.2
                reasons.append(f"fecha cercana ({date_diff} días)")
            elif date_diff <= 7:
                confidence += 0.1
                reasons.append(f"fecha similar ({date_diff} días)")
            
            if confidence > best_confidence:
                best_match = line
                best_confidence = confidence
                best_reason = ", ".join(reasons) if reasons else "coincidencia débil"
        
        if best_match and best_confidence >= 0.5:  # Solo sugerir si confianza >= 50%
            suggestions.append(MatchSuggestion(
                bank_transaction_id=tx.id,
                entry_line_id=best_match.id,
                confidence=best_confidence,
                reason=best_reason
            ))
    
    return suggestions

@router.post("/match")
def create_match(
    payload: MatchRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Crea un match entre una transacción bancaria y una línea contable"""
    if current_user.role not in (UserRole.ADMINISTRADOR, UserRole.CONTADOR):
        raise HTTPException(403, "No autorizado")
    
    # Verificar transacción bancaria
    bank_tx = db.query(BankTransaction).filter(BankTransaction.id == payload.bank_transaction_id).first()
    if not bank_tx:
        raise HTTPException(404, "Transacción bancaria no encontrada")
    
    if bank_tx.reconciled:
        raise HTTPException(400, "La transacción ya está conciliada")
    
    # Verificar línea contable
    entry_line = db.query(EntryLine).filter(EntryLine.id == payload.entry_line_id).first()
    if not entry_line:
        raise HTTPException(404, "Línea contable no encontrada")
    
    # Verificar que el monto coincida (aproximadamente)
    tx_amount = float(bank_tx.debit) if bank_tx.debit > 0 else float(bank_tx.credit)
    line_amount = float(entry_line.debit) if entry_line.debit > 0 else float(entry_line.credit)
    amount_diff = abs(tx_amount - line_amount)
    
    # Permitir conciliación aunque los montos no coincidan exactamente (puede haber diferencias por redondeo o ajustes)
    # Pero registrar la diferencia para auditoría
    if amount_diff > 0.01:
        # Log de advertencia pero permitir la conciliación
        import logging
        logger = logging.getLogger(__name__)
        logger.warning(
            f"Conciliación con diferencia de montos: "
            f"Transacción {bank_tx.id} (${tx_amount:.2f}) vs Línea {entry_line.id} (${line_amount:.2f}), "
            f"Diferencia: ${amount_diff:.2f}"
        )
    
    # Actualizar transacción
    bank_tx.entry_line_id = payload.entry_line_id
    bank_tx.reconciled = True
    
    db.commit()
    
    return {"success": True, "message": "Match creado exitosamente"}

@router.post("/bulk-match")
def create_bulk_matches(
    payload: BulkMatchRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Crea múltiples matches a la vez"""
    if current_user.role not in (UserRole.ADMINISTRADOR, UserRole.CONTADOR):
        raise HTTPException(403, "No autorizado")
    
    success_count = 0
    errors = []
    
    for match in payload.matches:
        try:
            create_match(match, db, current_user)
            success_count += 1
        except HTTPException as e:
            errors.append(f"Match {match.bank_transaction_id}-{match.entry_line_id}: {e.detail}")
    
    return {
        "success": len(errors) == 0,
        "success_count": success_count,
        "error_count": len(errors),
        "errors": errors
    }

@router.delete("/match/{bank_transaction_id}")
def remove_match(
    bank_transaction_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Elimina un match (desconciliar una transacción) - Revierte la conciliación"""
    if current_user.role not in (UserRole.ADMINISTRADOR, UserRole.CONTADOR):
        raise HTTPException(403, "No autorizado")
    
    bank_tx = db.query(BankTransaction).filter(BankTransaction.id == bank_transaction_id).first()
    if not bank_tx:
        raise HTTPException(404, "Transacción bancaria no encontrada")
    
    if not bank_tx.reconciled:
        raise HTTPException(400, "La transacción no está conciliada")
    
    # Deshacer conciliación
    bank_tx.entry_line_id = None
    bank_tx.reconciled = False
    
    db.commit()
    
    return {"success": True, "message": "Conciliación revertida exitosamente"}

# ===== HISTORIAL Y CONSULTA DE CONCILIACIONES =====

class ReconciledMatchOut(BaseModel):
    """Información de un match conciliado"""
    bank_transaction_id: int
    transaction_date: date
    transaction_description: str
    transaction_reference: str | None
    transaction_amount: float
    transaction_type: str  # "debit" o "credit"
    entry_line_id: int
    entry_date: date
    entry_glosa: str
    entry_memo: str | None
    entry_amount: float
    entry_type: str  # "debit" o "credit"
    amount_difference: float
    entry_number: str | None = None
    
    class Config:
        from_attributes = True

@router.get("/reconciled-matches/{bank_account_id}", response_model=List[ReconciledMatchOut])
def list_reconciled_matches(
    bank_account_id: int,
    period_id: int | None = None,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Lista todas las conciliaciones realizadas para una cuenta bancaria"""
    if current_user.role not in (UserRole.ADMINISTRADOR, UserRole.CONTADOR):
        raise HTTPException(403, "No autorizado")
    
    # Verificar cuenta bancaria
    bank_account = db.query(BankAccount).filter(BankAccount.id == bank_account_id).first()
    if not bank_account:
        raise HTTPException(404, "Cuenta bancaria no encontrada")
    
    # Obtener transacciones conciliadas
    query = (
        db.query(BankTransaction, EntryLine, JournalEntry)
        .join(EntryLine, BankTransaction.entry_line_id == EntryLine.id)
        .join(JournalEntry, EntryLine.entry_id == JournalEntry.id)
        .join(BankStatement, BankTransaction.statement_id == BankStatement.id)
        .filter(BankStatement.bank_account_id == bank_account_id)
        .filter(BankTransaction.reconciled == True)
    )
    
    if period_id:
        query = query.filter(BankStatement.period_id == period_id)
    
    results = query.order_by(BankTransaction.transaction_date.desc()).all()
    
    matches = []
    for bank_tx, entry_line, journal_entry in results:
        tx_amount = float(bank_tx.debit) if bank_tx.debit > 0 else float(bank_tx.credit)
        line_amount = float(entry_line.debit) if entry_line.debit > 0 else float(entry_line.credit)
        amount_diff = abs(tx_amount - line_amount)
        
        matches.append(ReconciledMatchOut(
            bank_transaction_id=bank_tx.id,
            transaction_date=bank_tx.transaction_date,
            transaction_description=bank_tx.description,
            transaction_reference=bank_tx.reference,
            transaction_amount=tx_amount,
            transaction_type="debit" if bank_tx.debit > 0 else "credit",
            entry_line_id=entry_line.id,
            entry_date=journal_entry.date,
            entry_glosa=journal_entry.glosa,
            entry_memo=entry_line.memo,
            entry_amount=line_amount,
            entry_type="debit" if entry_line.debit > 0 else "credit",
            amount_difference=amount_diff,
            entry_number=f"{journal_entry.id:06d}" if journal_entry else None
        ))
    
    return matches

@router.get("/reconciled-match-detail/{bank_transaction_id}")
def get_reconciled_match_detail(
    bank_transaction_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Obtiene el detalle completo de una conciliación específica"""
    if current_user.role not in (UserRole.ADMINISTRADOR, UserRole.CONTADOR):
        raise HTTPException(403, "No autorizado")
    
    result = (
        db.query(BankTransaction, EntryLine, JournalEntry, BankStatement, BankAccount, Account)
        .join(EntryLine, BankTransaction.entry_line_id == EntryLine.id)
        .join(JournalEntry, EntryLine.entry_id == JournalEntry.id)
        .join(BankStatement, BankTransaction.statement_id == BankStatement.id)
        .join(BankAccount, BankStatement.bank_account_id == BankAccount.id)
        .join(Account, EntryLine.account_id == Account.id)
        .filter(BankTransaction.id == bank_transaction_id)
        .filter(BankTransaction.reconciled == True)
        .first()
    )
    
    if not result:
        raise HTTPException(404, "Conciliación no encontrada")
    
    bank_tx, entry_line, journal_entry, statement, bank_account, account = result
    
    tx_amount = float(bank_tx.debit) if bank_tx.debit > 0 else float(bank_tx.credit)
    line_amount = float(entry_line.debit) if entry_line.debit > 0 else float(entry_line.credit)
    amount_diff = abs(tx_amount - line_amount)
    
    return {
        "bank_transaction": {
            "id": bank_tx.id,
            "date": bank_tx.transaction_date.isoformat(),
            "description": bank_tx.description,
            "reference": bank_tx.reference,
            "amount": tx_amount,
            "type": "debit" if bank_tx.debit > 0 else "credit",
            "balance": float(bank_tx.balance)
        },
        "entry_line": {
            "id": entry_line.id,
            "date": journal_entry.date.isoformat(),
            "glosa": journal_entry.glosa,
            "memo": entry_line.memo,
            "amount": line_amount,
            "type": "debit" if entry_line.debit > 0 else "credit",
            "account_code": account.code,
            "account_name": account.name
        },
        "journal_entry": {
            "id": journal_entry.id,
            "number": f"{journal_entry.id:06d}",
            "date": journal_entry.date.isoformat(),
            "glosa": journal_entry.glosa,
            "status": journal_entry.status
        },
        "statement": {
            "id": statement.id,
            "date": statement.statement_date.isoformat(),
            "period": f"{statement.period.year}-{statement.period.month:02d}" if statement.period else None
        },
        "bank_account": {
            "id": bank_account.id,
            "bank_name": bank_account.bank_name,
            "account_number": bank_account.account_number
        },
        "reconciliation": {
            "amount_difference": amount_diff,
            "amounts_match": amount_diff < 0.01,
            "has_warning": amount_diff > 0.01
        }
    }

@router.post("/finalize/{bank_account_id}")
def finalize_reconciliation(
    bank_account_id: int,
    period_id: int,
    payload: FinalizeReconciliationRequest,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """Finaliza la conciliación bancaria para un período"""
    if current_user.role not in (UserRole.ADMINISTRADOR, UserRole.CONTADOR):
        raise HTTPException(403, "No autorizado")
    
    bank_account = db.query(BankAccount).filter(BankAccount.id == bank_account_id).first()
    if not bank_account:
        raise HTTPException(404, "Cuenta bancaria no encontrada")
    
    period = db.query(Period).filter(Period.id == period_id).first()
    if not period:
        raise HTTPException(404, "Período no encontrado")
    
    # Obtener o crear conciliación
    reconciliation = db.query(BankReconciliation).filter(
        BankReconciliation.bank_account_id == bank_account_id,
        BankReconciliation.period_id == period_id
    ).first()
    
    statement = db.query(BankStatement).filter(
        BankStatement.bank_account_id == bank_account_id,
        BankStatement.period_id == period_id
    ).order_by(BankStatement.statement_date.desc()).first()
    
    # Calcular saldo contable
    from sqlalchemy import func
    
    all_periods = db.query(Period).filter(
        Period.company_id == period.company_id
    ).order_by(Period.year, Period.month).all()
    
    period_ids_list = [
        p.id for p in all_periods
        if (p.year < period.year) or (p.year == period.year and p.month <= period.month)
    ]
    
    book_balance_query = (
        db.query(func.coalesce(func.sum(EntryLine.debit) - func.sum(EntryLine.credit), Decimal('0')))
        .join(JournalEntry, JournalEntry.id == EntryLine.entry_id)
        .filter(EntryLine.account_id == bank_account.account_id)
        .filter(JournalEntry.period_id.in_(period_ids_list))
        .filter(JournalEntry.status == "POSTED")
    )
    book_balance = float(book_balance_query.scalar() or Decimal('0'))
    
    bank_balance = float(statement.closing_balance) if statement else 0.0
    
    # Validar que no haya líneas contables sin conciliar (opcional pero recomendado)
    # Obtener líneas contables del período que no están conciliadas
    unreconciled_lines = (
        db.query(EntryLine)
        .join(JournalEntry, JournalEntry.id == EntryLine.entry_id)
        .filter(EntryLine.account_id == bank_account.account_id)
        .filter(JournalEntry.period_id == period_id)
        .filter(JournalEntry.status == "POSTED")
        .filter(~EntryLine.id.in_(
            db.query(BankTransaction.entry_line_id)
            .join(BankStatement, BankStatement.id == BankTransaction.statement_id)
            .filter(BankStatement.bank_account_id == bank_account_id)
            .filter(BankStatement.period_id == period_id)
            .filter(BankTransaction.entry_line_id.isnot(None))
            .filter(BankTransaction.reconciled == True)
        ))
        .count()
    )
    
    # Advertencia si hay líneas sin conciliar (pero permitir finalizar)
    if unreconciled_lines > 0:
        # No bloqueamos, pero el usuario debería estar consciente
        pass  # Se puede agregar un warning en el response si se desea
    
    # Calcular saldo conciliado
    reconciled_balance = book_balance + float(payload.pending_credits) - float(payload.pending_debits)
    
    # Determinar estado
    difference = abs(reconciled_balance - bank_balance)
    if difference < 0.01:
        status = "CONCILIADO"
    elif difference < 100:  # Tolerancia de 100 unidades
        status = "PARCIAL"
    else:
        status = "PENDIENTE"
    
    if reconciliation:
        reconciliation.book_balance = Decimal(str(book_balance))
        reconciliation.bank_balance = Decimal(str(bank_balance))
        reconciliation.pending_debits = payload.pending_debits
        reconciliation.pending_credits = payload.pending_credits
        reconciliation.reconciled_balance = Decimal(str(reconciled_balance))
        reconciliation.status = status
        reconciliation.notes = payload.notes
        if status == "CONCILIADO":
            reconciliation.reconciled_at = datetime.now()
            reconciliation.reconciled_by = current_user.id
    else:
        reconciliation = BankReconciliation(
            bank_account_id=bank_account_id,
            period_id=period_id,
            statement_id=statement.id if statement else None,
            book_balance=Decimal(str(book_balance)),
            bank_balance=Decimal(str(bank_balance)),
            pending_debits=payload.pending_debits,
            pending_credits=payload.pending_credits,
            reconciled_balance=Decimal(str(reconciled_balance)),
            status=status,
            notes=payload.notes,
            reconciled_at=datetime.now() if status == "CONCILIADO" else None,
            reconciled_by=current_user.id if status == "CONCILIADO" else None
        )
        db.add(reconciliation)
    
    db.commit()
    db.refresh(reconciliation)
    
    return {
        "success": True,
        "status": status,
        "reconciled_balance": reconciled_balance,
        "bank_balance": bank_balance,
        "difference": difference,
        "unreconciled_lines_warning": unreconciled_lines if unreconciled_lines > 0 else None
    }

@router.post("/generate-test-data")
def generate_test_data(
    bank_account_id: int,
    period_id: int,
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Genera datos de prueba para conciliación bancaria:
    - Crea asientos contables que correspondan a transacciones bancarias
    - Crea un extracto bancario con transacciones
    """
    if current_user.role not in (UserRole.ADMINISTRADOR, UserRole.CONTADOR):
        raise HTTPException(403, "No autorizado")
    
    bank_account = db.query(BankAccount).filter(BankAccount.id == bank_account_id).first()
    if not bank_account:
        raise HTTPException(404, "Cuenta bancaria no encontrada")
    
    period = db.query(Period).filter(Period.id == period_id).first()
    if not period:
        raise HTTPException(404, "Período no encontrado")
    
    # Obtener cuenta contable
    account = db.query(Account).filter(Account.id == bank_account.account_id).first()
    if not account:
        raise HTTPException(404, "Cuenta contable no encontrada")
    
    # Generar transacciones de prueba
    transactions_data = [
        {"date": 5, "description": "Pago a proveedor ABC SAC", "reference": "CHQ-001", "amount": 1500.00, "type": "debit"},
        {"date": 8, "description": "Cobro de cliente XYZ EIRL", "reference": "TRF-001", "amount": 2500.00, "type": "credit"},
        {"date": 12, "description": "Pago servicios básicos", "reference": "CHQ-002", "amount": 350.00, "type": "debit"},
        {"date": 15, "description": "Depósito en efectivo", "reference": None, "amount": 800.00, "type": "credit"},
        {"date": 20, "description": "Transferencia a cuenta corriente", "reference": "TRF-002", "amount": 1200.00, "type": "debit"},
        {"date": 25, "description": "Cobro factura 001-001", "reference": "TRF-003", "amount": 1800.00, "type": "credit"},
    ]
    
    # Calcular saldo inicial (asumir 10000)
    opening_balance = Decimal('10000.00')
    current_balance = opening_balance
    
    # Crear asientos contables primero
    created_entries = []
    for tx_data in transactions_data:
        tx_date = date(period.year, period.month, tx_data["date"])
        amount = Decimal(str(tx_data["amount"]))
        
        # Crear asiento contable
        if tx_data["type"] == "debit":
            # Pago: Debe en cuenta bancaria, Haber en otra cuenta
            # Buscar cuenta de gastos o proveedores
            other_account = db.query(Account).filter(
                Account.company_id == bank_account.company_id,
                Account.code.like('60%')  # Gastos
            ).first()
            if not other_account:
                other_account = db.query(Account).filter(
                    Account.company_id == bank_account.company_id,
                    Account.code.like('42%')  # Proveedores
                ).first()
            
            if other_account:
                entry = JournalEntry(
                    company_id=bank_account.company_id,
                    date=tx_date,
                    period_id=period_id,
                    glosa=tx_data["description"],
                    currency=bank_account.currency,
                    exchange_rate=Decimal('1'),
                    origin="MANUAL",
                    status="POSTED"
                )
                db.add(entry)
                db.flush()
                
                # Línea 1: Debe en banco
                line1 = EntryLine(
                    entry_id=entry.id,
                    account_id=account.id,
                    debit=amount,
                    credit=Decimal('0'),
                    memo=tx_data["reference"]
                )
                # Línea 2: Haber en otra cuenta
                line2 = EntryLine(
                    entry_id=entry.id,
                    account_id=other_account.id,
                    debit=Decimal('0'),
                    credit=amount,
                    memo=tx_data["description"]
                )
                db.add(line1)
                db.add(line2)
                created_entries.append(entry)
        else:
            # Cobro: Haber en cuenta bancaria, Debe en otra cuenta
            # Buscar cuenta de ingresos o clientes
            other_account = db.query(Account).filter(
                Account.company_id == bank_account.company_id,
                Account.code.like('70%')  # Ingresos
            ).first()
            if not other_account:
                other_account = db.query(Account).filter(
                    Account.company_id == bank_account.company_id,
                    Account.code.like('12%')  # Clientes
                ).first()
            
            if other_account:
                entry = JournalEntry(
                    company_id=bank_account.company_id,
                    date=tx_date,
                    period_id=period_id,
                    glosa=tx_data["description"],
                    currency=bank_account.currency,
                    exchange_rate=Decimal('1'),
                    origin="MANUAL",
                    status="POSTED"
                )
                db.add(entry)
                db.flush()
                
                # Línea 1: Debe en otra cuenta
                line1 = EntryLine(
                    entry_id=entry.id,
                    account_id=other_account.id,
                    debit=amount,
                    credit=Decimal('0'),
                    memo=tx_data["description"]
                )
                # Línea 2: Haber en banco
                line2 = EntryLine(
                    entry_id=entry.id,
                    account_id=account.id,
                    debit=Decimal('0'),
                    credit=amount,
                    memo=tx_data["reference"]
                )
                db.add(line1)
                db.add(line2)
                created_entries.append(entry)
        
        # Actualizar saldo
        if tx_data["type"] == "debit":
            current_balance -= amount
        else:
            current_balance += amount
    
    db.commit()
    
    # Crear extracto bancario
    statement_date = date(period.year, period.month, 28)
    closing_balance = current_balance
    
    statement = BankStatement(
        bank_account_id=bank_account_id,
        period_id=period_id,
        statement_date=statement_date,
        opening_balance=opening_balance,
        closing_balance=closing_balance,
        uploaded_by=current_user.id,
        status="PENDIENTE"
    )
    db.add(statement)
    db.flush()
    
    # Crear transacciones bancarias
    current_tx_balance = opening_balance
    for tx_data in transactions_data:
        tx_date = date(period.year, period.month, tx_data["date"])
        amount = Decimal(str(tx_data["amount"]))
        
        if tx_data["type"] == "debit":
            current_tx_balance -= amount
            bank_tx = BankTransaction(
                statement_id=statement.id,
                transaction_date=tx_date,
                description=tx_data["description"],
                reference=tx_data["reference"],
                debit=amount,
                credit=Decimal('0'),
                balance=current_tx_balance,
                reconciled=False
            )
        else:
            current_tx_balance += amount
            bank_tx = BankTransaction(
                statement_id=statement.id,
                transaction_date=tx_date,
                description=tx_data["description"],
                reference=tx_data["reference"],
                debit=Decimal('0'),
                credit=amount,
                balance=current_tx_balance,
                reconciled=False
            )
        db.add(bank_tx)
    
    db.commit()
    db.refresh(statement)
    
    return {
        "success": True,
        "message": f"Datos de prueba generados exitosamente",
        "statement_id": statement.id,
        "transactions_count": len(transactions_data),
        "entries_created": len(created_entries),
        "opening_balance": float(opening_balance),
        "closing_balance": float(closing_balance)
    }


@router.get("/export/excel")
def export_reconciliation_excel(
    bank_account_id: int = Query(..., description="ID de la cuenta bancaria"),
    period_id: int = Query(..., description="ID del período"),
    db: Session = Depends(get_db),
    current_user: User = Depends(get_current_user)
):
    """
    Exporta el reporte de conciliación bancaria a Excel.
    Incluye: resumen, transacciones conciliadas, transacciones pendientes.
    """
    if current_user.role not in (UserRole.ADMINISTRADOR, UserRole.CONTADOR):
        raise HTTPException(403, "No autorizado")

    bank_account = db.query(BankAccount).filter(BankAccount.id == bank_account_id).first()
    if not bank_account:
        raise HTTPException(404, "Cuenta bancaria no encontrada")

    period = db.query(Period).filter(Period.id == period_id).first()
    if not period:
        raise HTTPException(404, "Período no encontrado")

    account = db.query(Account).filter(Account.id == bank_account.account_id).first()
    summary = get_reconciliation_summary(bank_account_id, period_id, db, current_user)

    # Obtener matches conciliados
    matches = list_reconciled_matches(bank_account_id, period_id, db, current_user)

    # Obtener transacciones no conciliadas
    transactions = get_unreconciled_transactions(bank_account_id, period_id, db, current_user)

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill
        from io import BytesIO

        wb = Workbook()
        ws = wb.active
        ws.title = "Conciliación"

        header_font = Font(bold=True)
        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font_white = Font(bold=True, color="FFFFFF")

        # Título
        ws.merge_cells("A1:F1")
        ws["A1"] = f"Reporte de Conciliación Bancaria - {bank_account.bank_name} {bank_account.account_number}"
        ws["A1"].font = Font(bold=True, size=14)
        ws["A2"] = f"Período: {period.year}-{period.month:02d}"
        ws["A3"] = f"Cuenta contable: {account.code if account else '-'} {account.name if account else ''}"

        # Resumen
        row = 5
        ws[f"A{row}"] = "RESUMEN"
        ws[f"A{row}"].font = header_font
        row += 1
        ws[f"A{row}"] = "Saldo según contabilidad"
        ws[f"B{row}"] = float(summary.book_balance)
        row += 1
        ws[f"A{row}"] = "Saldo según banco"
        ws[f"B{row}"] = float(summary.bank_balance)
        row += 1
        ws[f"A{row}"] = "Cheques pendientes"
        ws[f"B{row}"] = float(summary.pending_debits)
        row += 1
        ws[f"A{row}"] = "Depósitos en tránsito"
        ws[f"B{row}"] = float(summary.pending_credits)
        row += 1
        ws[f"A{row}"] = "Saldo conciliado"
        ws[f"B{row}"] = float(summary.reconciled_balance)
        row += 2

        # Transacciones conciliadas
        ws[f"A{row}"] = "TRANSACCIONES CONCILIADAS"
        ws[f"A{row}"].font = header_font
        row += 1
        headers = ["Fecha TX", "Descripción", "Monto TX", "Fecha Asiento", "Glosa", "Monto Asiento", "Diferencia"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = header_font_white
            cell.fill = header_fill
        row += 1
        for m in matches:
            ws.cell(row=row, column=1, value=m.transaction_date.isoformat())
            ws.cell(row=row, column=2, value=m.transaction_description or "")
            ws.cell(row=row, column=3, value=m.transaction_amount)
            ws.cell(row=row, column=4, value=m.entry_date.isoformat())
            ws.cell(row=row, column=5, value=m.entry_glosa or "")
            ws.cell(row=row, column=6, value=m.entry_amount)
            ws.cell(row=row, column=7, value=m.amount_difference)
            row += 1
        row += 1

        # Transacciones pendientes
        if transactions:
            ws[f"A{row}"] = "TRANSACCIONES PENDIENTES DE CONCILIAR"
            ws[f"A{row}"].font = header_font
            row += 1
            headers2 = ["Fecha", "Descripción", "Referencia", "Débito", "Crédito", "Saldo"]
            for col, h in enumerate(headers2, 1):
                cell = ws.cell(row=row, column=col, value=h)
                cell.font = header_font_white
                cell.fill = header_fill
            row += 1
            for tx in transactions:
                ws.cell(row=row, column=1, value=tx.transaction_date.isoformat())
                ws.cell(row=row, column=2, value=tx.description or "")
                ws.cell(row=row, column=3, value=tx.reference or "")
                ws.cell(row=row, column=4, value=float(tx.debit))
                ws.cell(row=row, column=5, value=float(tx.credit))
                ws.cell(row=row, column=6, value=float(tx.balance))
                row += 1

        buffer = BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        buffer_content = buffer.getvalue()
        filename = f"conciliacion_{bank_account.bank_name.replace(' ', '_')}_{period.year}{period.month:02d}.xlsx"

        return Response(
            content=buffer_content,
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={
                "Content-Disposition": f'attachment; filename="{filename}"',
                "Content-Length": str(len(buffer_content))
            }
        )
    except Exception as e:
        import logging
        logging.getLogger(__name__).error(f"Error exportando conciliación: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error al exportar: {str(e)}")
