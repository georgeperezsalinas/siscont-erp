"""
Extractores de Metadatos de Documentos
Extrae información estructurada de diferentes formatos de archivo
"""
from typing import Dict, Any, Optional
import re
from datetime import datetime
import io

try:
    import pdfplumber
    PDFPLUMBER_AVAILABLE = True
except ImportError:
    PDFPLUMBER_AVAILABLE = False

try:
    import openpyxl
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from xml.etree import ElementTree as ET
    XML_AVAILABLE = True
except ImportError:
    XML_AVAILABLE = False


class DocumentExtractor:
    """Extrae datos estructurados de documentos"""
    
    def extract_from_pdf(self, content: bytes) -> Dict[str, Any]:
        """
        Extrae datos de PDF (facturas, comprobantes).
        
        Retorna diccionario con:
        - ruc: RUC encontrado
        - dni: DNI encontrado
        - fecha: Fecha encontrada
        - total: Monto total
        - tipo_comprobante: Tipo de comprobante
        - serie: Serie del comprobante
        - numero: Número del comprobante
        """
        extracted = {}
        
        if not PDFPLUMBER_AVAILABLE:
            return extracted
        
        try:
            with pdfplumber.open(io.BytesIO(content)) as pdf:
                text = ""
                for page in pdf.pages:
                    page_text = page.extract_text()
                    if page_text:
                        text += page_text + "\n"
                
                # Extraer RUC (11 dígitos)
                ruc_patterns = [
                    r'RUC[:\s]*(\d{11})',
                    r'R\.U\.C\.\s*[:\s]*(\d{11})',
                    r'R\.U\.C[:\s]*(\d{11})',
                ]
                for pattern in ruc_patterns:
                    ruc_match = re.search(pattern, text, re.IGNORECASE)
                    if ruc_match:
                        extracted['ruc'] = ruc_match.group(1)
                        break
                
                # Extraer DNI (8 dígitos)
                dni_patterns = [
                    r'DNI[:\s]*(\d{8})',
                    r'D\.N\.I\.\s*[:\s]*(\d{8})',
                ]
                for pattern in dni_patterns:
                    dni_match = re.search(pattern, text, re.IGNORECASE)
                    if dni_match:
                        extracted['dni'] = dni_match.group(1)
                        break
                
                # Extraer fecha (múltiples formatos)
                date_patterns = [
                    r'(\d{2}/\d{2}/\d{4})',  # DD/MM/YYYY
                    r'(\d{4}-\d{2}-\d{2})',  # YYYY-MM-DD
                    r'(\d{2}-\d{2}-\d{4})',  # DD-MM-YYYY
                ]
                for pattern in date_patterns:
                    date_match = re.search(pattern, text)
                    if date_match:
                        extracted['fecha'] = date_match.group(1)
                        break
                
                # Extraer monto total
                total_patterns = [
                    r'TOTAL[:\s]*S/\.?\s*([\d,]+\.?\d*)',
                    r'IMPORTE TOTAL[:\s]*S/\.?\s*([\d,]+\.?\d*)',
                    r'TOTAL A PAGAR[:\s]*S/\.?\s*([\d,]+\.?\d*)',
                    r'MONTO TOTAL[:\s]*S/\.?\s*([\d,]+\.?\d*)',
                ]
                for pattern in total_patterns:
                    total_match = re.search(pattern, text, re.IGNORECASE)
                    if total_match:
                        try:
                            total_str = total_match.group(1).replace(',', '')
                            extracted['total'] = float(total_str)
                        except ValueError:
                            pass
                        break
                
                # Extraer número de comprobante
                comprobante_patterns = [
                    r'(FACTURA|BOLETA|NOTA\s+CREDITO|NOTA\s+DEBITO)[:\s]*(\w+)[-\s]*(\d+)',
                    r'(\w+)[-\s]*(\d+)\s*(FACTURA|BOLETA|NOTA)',
                ]
                for pattern in comprobante_patterns:
                    comprobante_match = re.search(pattern, text, re.IGNORECASE)
                    if comprobante_match:
                        if len(comprobante_match.groups()) >= 3:
                            extracted['tipo_comprobante'] = comprobante_match.group(1).strip()
                            extracted['serie'] = comprobante_match.group(2).strip()
                            extracted['numero'] = comprobante_match.group(3).strip()
                        break
        
        except Exception as e:
            # Log error pero no fallar
            print(f"Error extrayendo datos de PDF: {e}")
        
        return extracted
    
    def extract_from_xml(self, content: bytes) -> Dict[str, Any]:
        """
        Extrae datos de XML (factura electrónica SUNAT).
        
        Retorna diccionario con datos estructurados del XML.
        """
        extracted = {}
        
        if not XML_AVAILABLE:
            return extracted
        
        try:
            root = ET.fromstring(content)
            
            # Namespaces comunes en facturas electrónicas UBL 2.1
            namespaces = {
                'cac': 'urn:oasis:names:specification:ubl:schema:xsd:CommonAggregateComponents-2',
                'cbc': 'urn:oasis:names:specification:ubl:schema:xsd:CommonBasicComponents-2',
                'ext': 'urn:oasis:names:specification:ubl:schema:xsd:CommonExtensionComponents-2'
            }
            
            # Intentar sin namespace primero (algunos XMLs no los usan)
            try:
                # RUC emisor
                ruc_elem = root.find('.//cac:PartyIdentification/cbc:ID', namespaces)
                if ruc_elem is None:
                    ruc_elem = root.find('.//PartyIdentification/ID')
                if ruc_elem is not None:
                    extracted['ruc_emisor'] = ruc_elem.text
                
                # Fecha de emisión
                fecha_elem = root.find('.//cbc:IssueDate', namespaces)
                if fecha_elem is None:
                    fecha_elem = root.find('.//IssueDate')
                if fecha_elem is not None:
                    extracted['fecha'] = fecha_elem.text
                
                # Número de comprobante
                numero_elem = root.find('.//cbc:ID', namespaces)
                if numero_elem is None:
                    numero_elem = root.find('.//ID')
                if numero_elem is not None:
                    extracted['numero_comprobante'] = numero_elem.text
                
                # Monto total
                total_elem = root.find('.//cac:LegalMonetaryTotal/cbc:PayableAmount', namespaces)
                if total_elem is None:
                    total_elem = root.find('.//LegalMonetaryTotal/PayableAmount')
                if total_elem is not None:
                    try:
                        extracted['total'] = float(total_elem.text)
                    except (ValueError, AttributeError):
                        pass
                
                # Tipo de documento
                doc_type_elem = root.find('.//cbc:InvoiceTypeCode', namespaces)
                if doc_type_elem is None:
                    doc_type_elem = root.find('.//InvoiceTypeCode')
                if doc_type_elem is not None:
                    extracted['tipo_documento'] = doc_type_elem.text
                
            except Exception:
                # Si falla con namespaces, intentar búsqueda simple
                pass
        
        except Exception as e:
            print(f"Error extrayendo datos de XML: {e}")
        
        return extracted
    
    def extract_from_excel(self, content: bytes) -> Dict[str, Any]:
        """
        Extrae datos básicos de Excel.
        
        Retorna diccionario con:
        - headers: Lista de headers de la primera fila
        - row_count: Número de filas de datos
        - sheet_names: Lista de nombres de hojas
        """
        extracted = {}
        
        if not OPENPYXL_AVAILABLE:
            return extracted
        
        try:
            workbook = openpyxl.load_workbook(io.BytesIO(content))
            
            # Información de hojas
            extracted['sheet_names'] = workbook.sheetnames
            
            # Procesar primera hoja
            if workbook.sheetnames:
                sheet = workbook[workbook.sheetnames[0]]
                
                # Headers de la primera fila
                headers = []
                if sheet.max_row > 0:
                    for cell in sheet[1]:
                        if cell.value:
                            headers.append(str(cell.value))
                
                extracted['headers'] = headers
                extracted['row_count'] = sheet.max_row - 1 if sheet.max_row > 1 else 0
        
        except Exception as e:
            print(f"Error extrayendo datos de Excel: {e}")
        
        return extracted
    
    def extract_from_text(self, content: bytes, encoding: str = 'utf-8') -> Dict[str, Any]:
        """
        Extrae datos básicos de archivo de texto.
        
        Retorna diccionario con:
        - text_preview: Primeras 500 caracteres
        - line_count: Número de líneas
        """
        extracted = {}
        
        try:
            text = content.decode(encoding)
            lines = text.split('\n')
            
            extracted['text_preview'] = text[:500]
            extracted['line_count'] = len(lines)
            extracted['char_count'] = len(text)
        
        except Exception as e:
            print(f"Error extrayendo datos de texto: {e}")
        
        return extracted

